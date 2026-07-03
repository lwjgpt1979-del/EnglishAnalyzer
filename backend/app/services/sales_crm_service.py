"""电销 CRM 服务(P0):线索池 + 公海/私海 + 跟进 + 赢单画像反查推荐。

外呼/ASR/意向分析(P1)、企微存档(P2)后续接入;本层只做零第三方依赖的闭环。
运营可配置值(公海回收天数、推荐权重等)读 system_configs.sales_crm,禁写死。
方案见 docs/电销CRM-方案设计.md。
"""
from __future__ import annotations

import uuid
from datetime import datetime, timedelta, timezone

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AppError
from app.models.d9_system import SystemConfig
from app.models.d23_sales_crm import SalesLead, SalesLeadActivity
from app.services import region_service

_CFG_KEY = "sales_crm"

# 常量仅兜底/白名单;实际值见 get_config()(system_configs.sales_crm)。
DEFAULTS: dict = {
    "public_pool_recycle_days": 7,      # 私海 N 天未跟进 → 回收公海
    "recommend_weights": {              # 赢单画像反查各维度权重
        "industry": 3.0, "province": 1.0, "city": 2.0, "tag": 1.0,
    },
    "intent_grade_thresholds": {        # 意向分 → 分层(A/B/C,其余 D)
        "A": 80, "B": 60, "C": 40,
    },
    "sla_overdue_hours": 48,            # 跟进 SLA:next_follow_at 超时超过 N 小时 → 违约告警
    "seat_only_admin_ids": [],          # 座席名单(id 字符串):名单内只看公海+自己私海;名单外看全部
    "tag_catalog": ["高意向", "价格敏感", "已加微信", "决策人", "需回访", "同行"],  # 运营标签建议
}

STATUSES = ("new", "contacted", "interested", "negotiating", "won", "lost", "invalid")
SOURCES = ("baidu_map", "amap", "meituan", "dianping", "tungee", "manual", "import", "other")
CHANNELS = ("call", "wechat", "note", "sms")


async def get_config(db: AsyncSession) -> dict:
    row = (await db.execute(
        sa.select(SystemConfig).where(SystemConfig.key == _CFG_KEY))).scalar_one_or_none()
    cfg = dict(DEFAULTS)
    if row is not None and isinstance(row.value, dict):
        for k in DEFAULTS:
            if k in row.value:
                cfg[k] = row.value[k]
    return cfg


async def update_config(db: AsyncSession, *, patch: dict, updated_by: uuid.UUID) -> dict:
    clean = {k: v for k, v in (patch or {}).items() if k in DEFAULTS}
    row = (await db.execute(
        sa.select(SystemConfig).where(SystemConfig.key == _CFG_KEY))).scalar_one_or_none()
    merged = dict(DEFAULTS)
    if row is not None and isinstance(row.value, dict):
        merged.update(row.value)
    merged.update(clean)
    if row is None:
        db.add(SystemConfig(id=uuid.uuid4(), key=_CFG_KEY, value=merged,
                            description="电销 CRM 参数(公海回收天数/推荐权重)", updated_by=updated_by))
    else:
        row.value = merged
        row.updated_by = updated_by
    await db.flush()
    return {k: merged.get(k, DEFAULTS[k]) for k in DEFAULTS}


# ── 线索 CRUD / 池 ────────────────────────────────────────────────────────────

async def _resolve_region(db: AsyncSession, *, region_code: str | None,
                          region_name: str | None) -> tuple[str | None, str | None]:
    """地区一律走 region_service:有码用码取名;只有名字则 region_from_name 反解。"""
    if region_code:
        return region_code, region_name
    if region_name:
        code, name = await region_service.region_from_name(db, region_name)
        if code:
            return code, name or region_name
    return None, region_name


async def create_lead(db: AsyncSession, *, data: dict) -> SalesLead:
    rc, rn = await _resolve_region(
        db, region_code=data.get("region_code"), region_name=data.get("region_name"))
    lead = SalesLead(
        id=uuid.uuid4(),
        name=(data.get("name") or "").strip() or "未命名线索",
        contact_name=data.get("contact_name"),
        phone=(data.get("phone") or "").strip() or None,
        wechat_id=data.get("wechat_id"),
        address=data.get("address"),
        region_code=rc, region_name=rn,
        industry=data.get("industry"),
        biz_tags=data.get("biz_tags"),
        tags=data.get("tags"),
        source=data.get("source") if data.get("source") in SOURCES else "manual",
        source_note=data.get("source_note"),
        status="new",
        consent=bool(data.get("consent", False)),
        dnc=bool(data.get("dnc", False)),
        pool="public",
    )
    db.add(lead)
    await db.flush()
    return lead


import re as _re

# 多号分隔符:仅逗号/分号/顿号/斜杠(不含空格——单个号码内部可能带空格,不能据此截断)
_PHONE_SPLIT = _re.compile(r"[,，;；/、]+")


def _first_phone(raw: object) -> str | None:
    """一个字段里可能有多号(「025-83xxxxxx,138xxxx」)→ 取首个 ≥7 位数字的号,
    清掉空格/括号等杂字符(只留数字与短横),保住区号-号码完整。"""
    if not raw:
        return None
    for part in _PHONE_SPLIT.split(str(raw)):
        cleaned = _re.sub(r"[^\d\-]", "", part).strip("-")
        if len(_re.sub(r"\D", "", cleaned)) >= 7:
            return cleaned or None
    return None


_SHARED_TAG = "同号多机构"   # 同一电话挂在不同地址 → 疑似一个老板多店/多机构
_NOPHONE_TAG = "待补号"      # 无电话线索:入库但需人工补号才能外呼


def _merge_tags(tags: object, tag: str) -> list:
    lst = list(tags) if isinstance(tags, list) else []
    if tag not in lst:
        lst.append(tag)
    return lst


def _norm_addr(a: object) -> str:
    return _re.sub(r"\s+", "", (str(a).strip() if a else ""))


async def ingest_external_leads(db: AsyncSession, *, items: list[dict], source: str,
                                source_note: str | None = None,
                                require_phone: bool = False) -> dict:
    """采集→入库通用适配器:{name,phone,address,business,city} 列表 → 经 region_service 解析城市/地址
    为 region_code → 归一 → 智能去重入库。source 自动填(非法值兜底 other),source_note 记来源(合规)。
    不管数据来自百度 API / 探迹 / Excel,都过这一层。

    去重:
    - 有电话:同号同址=真重复跳过;同号不同址=保留 + 打「同号多机构」标(新旧线索都标,疑似一个老板多店)。
    - 无电话:按「同名 + 地址」去重——同名同址跳过;同名不同址照样入库(不打同号标)。
    require_phone=True 时无电话直接跳过(默认 False:无电话也入库)。
    返回 {created, skipped, shared_phone, no_phone, region_unresolved}。
    """
    from app.services import region_service
    src = source if source in SOURCES else "other"
    norm, no_phone, unresolved = [], 0, 0
    for it in items:
        phone = _first_phone(it.get("phone"))
        if not phone:
            no_phone += 1
            if require_phone:
                continue
        city = (it.get("city") or "").strip()
        addr = (it.get("address") or "").strip()
        code = rname = None
        if city:
            code, rname = await region_service.region_from_name(db, city)
        if code is None and addr:
            code, rname = await region_service.region_from_name(db, addr)
        if code is None:
            unresolved += 1
        norm.append({
            "name": (it.get("name") or "").strip() or "未命名机构",
            "phone": phone,
            "address": addr or None,
            "region_code": code, "region_name": rname,
            "industry": (it.get("business") or it.get("industry") or "").strip() or None,
            "source": src, "source_note": source_note,
        })

    # 预取已存在线索:有电话按 phone,无电话按 name(仅无电话的历史线索)
    phones = {n["phone"] for n in norm if n.get("phone")}
    names_np = {n["name"] for n in norm if not n.get("phone")}
    exist_ph: dict[str, list] = {}
    exist_nm: dict[str, list] = {}
    if phones:
        for lead in (await db.execute(sa.select(SalesLead).where(SalesLead.phone.in_(phones)))).scalars().all():
            exist_ph.setdefault(lead.phone, []).append(lead)
    if names_np:
        for lead in (await db.execute(sa.select(SalesLead).where(
                SalesLead.phone.is_(None), SalesLead.name.in_(names_np)))).scalars().all():
            exist_nm.setdefault(lead.name, []).append(lead)

    created = skipped = shared = 0
    batch_ph: dict[str, set] = {}
    batch_nm: dict[str, set] = {}
    flag_existing: set = set()
    for n in norm:
        phone, name, an = n.get("phone"), n["name"], _norm_addr(n.get("address"))
        if phone:
            seen = {_norm_addr(l.address) for l in exist_ph.get(phone, [])} | batch_ph.get(phone, set())
            if an in seen:               # 同号同址 → 真重复
                skipped += 1
                continue
            if exist_ph.get(phone) or batch_ph.get(phone):   # 同号不同址 → 一号多店
                n["tags"] = _merge_tags(n.get("tags"), _SHARED_TAG)
                shared += 1
                flag_existing.update(exist_ph.get(phone, []))
            batch_ph.setdefault(phone, set()).add(an)
        else:                            # 无电话:按 同名+地址 去重,并打「待补号」标
            seen = {_norm_addr(l.address) for l in exist_nm.get(name, [])} | batch_nm.get(name, set())
            if an in seen:               # 同名同址 → 真重复
                skipped += 1
                continue
            n["tags"] = _merge_tags(n.get("tags"), _NOPHONE_TAG)
            batch_nm.setdefault(name, set()).add(an)
        await create_lead(db, data=n)
        created += 1

    for lead in flag_existing:           # 已存在的同号线索也补标
        lead.tags = _merge_tags(lead.tags, _SHARED_TAG)
    await db.flush()

    return {"created": created, "skipped": skipped, "shared_phone": shared,
            "no_phone": no_phone, "region_unresolved": unresolved}


async def import_leads(db: AsyncSession, *, items: list[dict], source: str = "import") -> dict:
    """批量导入:按 phone 去重(已存在则跳过)。返回 {created, skipped}。"""
    created = skipped = 0
    phones = [(it.get("phone") or "").strip() for it in items if (it.get("phone") or "").strip()]
    existing: set[str] = set()
    if phones:
        existing = set((await db.execute(
            sa.select(SalesLead.phone).where(SalesLead.phone.in_(phones)))).scalars().all())
    for it in items:
        phone = (it.get("phone") or "").strip()
        if phone and phone in existing:
            skipped += 1
            continue
        it = dict(it)
        it.setdefault("source", source)
        await create_lead(db, data=it)
        if phone:
            existing.add(phone)
        created += 1
    return {"created": created, "skipped": skipped}


async def list_leads(
    db: AsyncSession, *, pool: str | None = None, status: str | None = None,
    source: str | None = None, region_code: str | None = None,
    owner_admin_id: uuid.UUID | None = None, dnc: bool | None = None,
    has_phone: bool | None = None,
    due: bool = False, sla: bool = False, tag: str | None = None,
    seat_admin_id: uuid.UUID | None = None, q: str | None = None,
    skip: int = 0, limit: int = 20,
) -> tuple[list[SalesLead], int]:
    base = sa.select(SalesLead)
    if seat_admin_id is not None:   # 座席权限:只看公海 + 自己私海
        base = base.where(sa.or_(SalesLead.pool == "public",
                                 SalesLead.owner_admin_id == seat_admin_id))
    if tag:
        base = base.where(SalesLead.tags.contains([tag]))
    if has_phone is True:           # 有电话
        base = base.where(SalesLead.phone.isnot(None))
    elif has_phone is False:        # 无电话(待补号)
        base = base.where(SalesLead.phone.is_(None))
    if pool:
        base = base.where(SalesLead.pool == pool)
    if status:
        base = base.where(SalesLead.status == status)
    if source:
        base = base.where(SalesLead.source == source)
    if region_code:               # 前缀:省码含其下所有市
        base = base.where(SalesLead.region_code.like(f"{region_code}%"))
    if owner_admin_id is not None:
        base = base.where(SalesLead.owner_admin_id == owner_admin_id)
    if dnc is not None:
        base = base.where(SalesLead.dnc.is_(dnc))
    if due:                       # 今日待办:已到跟进时间、且未到终态
        base = base.where(
            SalesLead.next_follow_at.isnot(None),
            SalesLead.next_follow_at <= datetime.now(timezone.utc),
            SalesLead.status.notin_(("won", "lost", "invalid")))
    if sla:                       # SLA 违约:超时超过阈值
        cfg = await get_config(db)
        cutoff = datetime.now(timezone.utc) - timedelta(hours=int(cfg["sla_overdue_hours"]))
        base = base.where(
            SalesLead.next_follow_at.isnot(None),
            SalesLead.next_follow_at < cutoff,
            SalesLead.status.notin_(("won", "lost", "invalid")))
    if q:
        like = f"%{q}%"
        base = base.where(sa.or_(SalesLead.name.ilike(like),
                                 SalesLead.phone.ilike(like),
                                 SalesLead.contact_name.ilike(like)))
    total = (await db.execute(
        sa.select(sa.func.count()).select_from(base.subquery()))).scalar_one()
    rows = (await db.execute(
        base.order_by(SalesLead.next_follow_at.asc().nullslast(),
                      SalesLead.created_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    return list(rows), total


async def get_lead(db: AsyncSession, lead_id: uuid.UUID) -> SalesLead:
    lead = await db.get(SalesLead, lead_id)
    if lead is None:
        raise AppError(code=404, message="线索不存在")
    return lead


_EDITABLE = {"name", "contact_name", "phone", "wechat_id", "address", "industry",
             "biz_tags", "tags", "source_note", "consent", "dnc", "next_follow_at",
             "intent_score", "intent_grade"}


async def update_lead(db: AsyncSession, *, lead_id: uuid.UUID, patch: dict) -> SalesLead:
    lead = await get_lead(db, lead_id)
    for k, v in (patch or {}).items():
        if k in _EDITABLE:
            setattr(lead, k, v)
    if "status" in (patch or {}) and patch["status"] in STATUSES:
        lead.status = patch["status"]
    if "region_code" in (patch or {}) or "region_name" in (patch or {}):
        rc, rn = await _resolve_region(
            db, region_code=patch.get("region_code"), region_name=patch.get("region_name"))
        lead.region_code, lead.region_name = rc, rn
    await db.flush()
    return lead


async def claim_lead(db: AsyncSession, *, lead_id: uuid.UUID, admin_id: uuid.UUID) -> SalesLead:
    """认领进私海。已被他人认领则拒绝(防撞单)。"""
    lead = await get_lead(db, lead_id)
    if lead.pool == "private" and lead.owner_admin_id not in (None, admin_id):
        raise AppError(code=409, message="该线索已被他人认领")
    lead.pool = "private"
    lead.owner_admin_id = admin_id
    lead.claimed_at = datetime.now(timezone.utc)
    await db.flush()
    return lead


async def release_lead(db: AsyncSession, *, lead_id: uuid.UUID) -> SalesLead:
    """退回公海。"""
    lead = await get_lead(db, lead_id)
    lead.pool = "public"
    lead.owner_admin_id = None
    lead.claimed_at = None
    await db.flush()
    return lead


async def batch_assign(
    db: AsyncSession, *, lead_ids: list[uuid.UUID], owner_admin_id: uuid.UUID,
) -> int:
    """批量派单/认领:把选中线索分配给某座席(→私海)。返回分配数。"""
    if not lead_ids:
        return 0
    r = await db.execute(
        sa.update(SalesLead).where(SalesLead.id.in_(lead_ids)).values(
            pool="private", owner_admin_id=owner_admin_id,
            claimed_at=datetime.now(timezone.utc)))
    await db.flush()
    return r.rowcount


async def list_seats(db: AsyncSession) -> list[dict]:
    """座席列表(平台管理员),供批量派单选人。"""
    from app.models.d1_users import User
    rows = (await db.execute(
        sa.select(User.id, User.nickname, User.username, User.phone)
        .where(User.role == "platform_admin").order_by(User.username))).all()
    return [{"id": str(uid), "name": (nick or uname or (phone or "")[-4:] or "管理员")}
            for uid, nick, uname, phone in rows]


async def recycle_public_pool(db: AsyncSession) -> int:
    """私海超 N 天未跟进(last_contacted_at / claimed_at 取新者为基准)→ 回收公海。返回回收数。"""
    cfg = await get_config(db)
    days = int(cfg["public_pool_recycle_days"])
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    stale = (await db.execute(sa.select(SalesLead).where(
        SalesLead.pool == "private",
        sa.func.coalesce(SalesLead.last_contacted_at, SalesLead.claimed_at) < cutoff,
        SalesLead.status.notin_(("won", "negotiating")),   # 谈单中/已成不回收
    ))).scalars().all()
    for lead in stale:
        lead.pool = "public"
        lead.owner_admin_id = None
        lead.claimed_at = None
    await db.flush()
    return len(stale)


# ── 跟进记录 ──────────────────────────────────────────────────────────────────

async def add_activity(
    db: AsyncSession, *, lead_id: uuid.UUID, admin_id: uuid.UUID | None,
    channel: str, content: str | None = None, direction: str | None = None,
    outcome: str | None = None, next_follow_at: datetime | None = None,
    status: str | None = None,
) -> SalesLeadActivity:
    lead = await get_lead(db, lead_id)
    if channel not in CHANNELS:
        raise AppError(code=400, message="非法跟进方式")
    act = SalesLeadActivity(
        id=uuid.uuid4(), lead_id=lead_id, admin_id=admin_id,
        channel=channel, direction=direction, content=content, outcome=outcome)
    db.add(act)
    lead.last_contacted_at = datetime.now(timezone.utc)
    if next_follow_at is not None:
        lead.next_follow_at = next_follow_at
    if status and status in STATUSES:
        lead.status = status
    elif lead.status == "new":       # 首次触达自动推进
        lead.status = "contacted"
    await db.flush()
    return act


async def list_activities(
    db: AsyncSession, *, lead_id: uuid.UUID, skip: int = 0, limit: int = 20,
) -> tuple[list[SalesLeadActivity], int]:
    base = sa.select(SalesLeadActivity).where(SalesLeadActivity.lead_id == lead_id)
    total = (await db.execute(
        sa.select(sa.func.count()).select_from(base.subquery()))).scalar_one()
    rows = (await db.execute(
        base.order_by(SalesLeadActivity.created_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    return list(rows), total


# ── 赢单画像反查推荐(P0 纯查询自研) ────────────────────────────────────────

_RECO_CANDIDATE_CAP = 2000    # 打分候选上限,超出按 next_follow/created 序截断(避免全表扫)


async def recommend(
    db: AsyncSession, *, skip: int = 0, limit: int = 20,
) -> tuple[list[SalesLead], int]:
    """探迹式相似推荐:用 status=won 线索画像(行业/省/市/经营标签)给公海新线索打分。

    返回按 similar_score 降序的公海 new 线索页;顺带把分写回 similar_score。
    """
    cfg = await get_config(db)
    w = {**DEFAULTS["recommend_weights"], **(cfg.get("recommend_weights") or {})}

    won = (await db.execute(sa.select(
        SalesLead.industry, SalesLead.region_code, SalesLead.biz_tags
    ).where(SalesLead.status == "won"))).all()

    ind_set: set[str] = set()
    prov_set: set[str] = set()
    city_set: set[str] = set()
    tag_set: set[str] = set()
    for ind, rc, tags in won:
        if ind:
            ind_set.add(ind)
        if rc:
            prov_set.add(rc[:2])
            if len(rc) >= 4:
                city_set.add(rc[:4])
        if isinstance(tags, list):
            tag_set.update(str(t) for t in tags)

    cands = (await db.execute(sa.select(SalesLead).where(
        SalesLead.pool == "public", SalesLead.status == "new"
    ).order_by(SalesLead.next_follow_at.asc().nullslast(),
               SalesLead.created_at.desc()).limit(_RECO_CANDIDATE_CAP))).scalars().all()

    scored: list[tuple[float, SalesLead]] = []
    for c in cands:
        s = 0.0
        if not won:
            s = 0.0                              # 无赢单样本 → 全 0,退化为普通公海列表
        else:
            if c.industry and c.industry in ind_set:
                s += w["industry"]
            if c.region_code:
                if c.region_code[:2] in prov_set:
                    s += w["province"]
                if len(c.region_code) >= 4 and c.region_code[:4] in city_set:
                    s += w["city"]
            if isinstance(c.biz_tags, list) and tag_set:
                overlap = len({str(t) for t in c.biz_tags} & tag_set)
                s += w["tag"] * overlap
        c.similar_score = s
        scored.append((s, c))

    scored.sort(key=lambda x: (-x[0], x[1].created_at and 0))
    await db.flush()
    total = len(scored)
    page = [c for _s, c in scored[skip:skip + limit]]
    return page, total


# ── 座席看板 ──────────────────────────────────────────────────────────────────

_CN_TZ = timezone(timedelta(hours=8))    # 「今日」按东八区(业务在国内)


def _today_start_utc() -> datetime:
    now_cn = datetime.now(_CN_TZ)
    return now_cn.replace(hour=0, minute=0, second=0, microsecond=0)


async def board_stats(db: AsyncSession, *, admin_id: uuid.UUID | None = None) -> dict:
    """座席看板:线索分布 + 今日拨打量/接通率/今日新增 + 我的待办数。"""
    by_status = dict((await db.execute(
        sa.select(SalesLead.status, sa.func.count()).group_by(SalesLead.status))).all())
    by_pool = dict((await db.execute(
        sa.select(SalesLead.pool, sa.func.count()).group_by(SalesLead.pool))).all())
    total = (await db.execute(sa.select(sa.func.count()).select_from(SalesLead))).scalar_one()

    today = _today_start_utc()
    now = datetime.now(timezone.utc)
    today_new = (await db.execute(sa.select(sa.func.count()).where(
        SalesLead.created_at >= today))).scalar_one()
    today_calls = (await db.execute(sa.select(sa.func.count()).where(
        SalesLeadActivity.channel == "call",
        SalesLeadActivity.created_at >= today))).scalar_one()
    today_connected = (await db.execute(sa.select(sa.func.count()).where(
        SalesLeadActivity.channel == "call",
        SalesLeadActivity.outcome == "connected",
        SalesLeadActivity.created_at >= today))).scalar_one()
    # 待办:已到跟进时间、未到终态(admin 指定则只算其私海)
    due_q = sa.select(sa.func.count()).where(
        SalesLead.next_follow_at.isnot(None), SalesLead.next_follow_at <= now,
        SalesLead.status.notin_(("won", "lost", "invalid")))
    if admin_id is not None:
        due_q = due_q.where(SalesLead.owner_admin_id == admin_id)
    my_due = (await db.execute(due_q)).scalar_one()

    cfg = await get_config(db)
    sla_cutoff = now - timedelta(hours=int(cfg["sla_overdue_hours"]))
    sla_breach = (await db.execute(sa.select(sa.func.count()).where(
        SalesLead.next_follow_at.isnot(None), SalesLead.next_follow_at < sla_cutoff,
        SalesLead.status.notin_(("won", "lost", "invalid"))))).scalar_one()

    return {
        "total": int(total),
        "by_status": {k: int(v) for k, v in by_status.items()},
        "by_pool": {k: int(v) for k, v in by_pool.items()},
        "today_new": int(today_new),
        "today_calls": int(today_calls),
        "today_connected": int(today_connected),
        "connect_rate": round(today_connected / today_calls, 3) if today_calls else 0.0,
        "my_due": int(my_due),
        "sla_breach": int(sla_breach),
        "sla_overdue_hours": int(cfg["sla_overdue_hours"]),
    }


# ── 来源统计 / 查重合并 ────────────────────────────────────────────────────────

async def source_stats(db: AsyncSession) -> list[dict]:
    """按来源统计:线索数 + 成交数 + 转化率。供来源看板。"""
    total_by = dict((await db.execute(
        sa.select(SalesLead.source, sa.func.count()).group_by(SalesLead.source))).all())
    won_by = dict((await db.execute(
        sa.select(SalesLead.source, sa.func.count())
        .where(SalesLead.status == "won").group_by(SalesLead.source))).all())
    out = []
    for src, cnt in sorted(total_by.items(), key=lambda x: -x[1]):
        won = int(won_by.get(src, 0))
        out.append({"source": src, "total": int(cnt), "won": won,
                    "conversion": round(won / cnt, 3) if cnt else 0.0})
    return out


async def find_duplicate_groups(db: AsyncSession, *, limit: int = 100) -> list[dict]:
    """按电话找重复线索组(同一非空 phone 有 ≥2 条)。返回 [{phone, leads:[...]}]。"""
    dup_phones = (await db.execute(
        sa.select(SalesLead.phone).where(SalesLead.phone.isnot(None), SalesLead.phone != "")
        .group_by(SalesLead.phone).having(sa.func.count() > 1).limit(limit))).scalars().all()
    if not dup_phones:
        return []
    rows = (await db.execute(sa.select(SalesLead).where(SalesLead.phone.in_(dup_phones))
                             .order_by(SalesLead.phone, SalesLead.created_at))).scalars().all()
    groups: dict[str, list] = {}
    for r in rows:
        groups.setdefault(r.phone, []).append(r)
    return [{"phone": ph, "leads": [_lead_brief(x) for x in ls]} for ph, ls in groups.items()]


def _lead_brief(l: SalesLead) -> dict:
    return {"id": str(l.id), "name": l.name, "contact_name": l.contact_name,
            "region_name": l.region_name, "status": l.status, "source": l.source,
            "pool": l.pool, "created_at": l.created_at.isoformat() if l.created_at else None}


async def merge_leads(db: AsyncSession, *, survivor_id: uuid.UUID,
                      dup_ids: list[uuid.UUID]) -> dict:
    """合并:把 dup 的跟进/企微记录改挂到 survivor,合并产品意见 + 补空字段,删 dup。"""
    from app.models.d23_sales_crm import WecomChatArchive
    dup_ids = [d for d in dup_ids if d != survivor_id]
    if not dup_ids:
        return {"merged": 0}
    survivor = await get_lead(db, survivor_id)
    dups = (await db.execute(
        sa.select(SalesLead).where(SalesLead.id.in_(dup_ids)))).scalars().all()
    moved_acts = (await db.execute(
        sa.update(SalesLeadActivity).where(SalesLeadActivity.lead_id.in_(dup_ids))
        .values(lead_id=survivor_id))).rowcount
    moved_wecom = (await db.execute(
        sa.update(WecomChatArchive).where(WecomChatArchive.lead_id.in_(dup_ids))
        .values(lead_id=survivor_id))).rowcount
    # 合并产品意见 + 补 survivor 的空字段
    fb = list(survivor.product_feedback or [])
    for d in dups:
        for f in (d.product_feedback or []):
            if f not in fb:
                fb.append(f)
        for attr in ("contact_name", "wechat_id", "address", "region_code",
                     "region_name", "industry", "source_note"):
            if not getattr(survivor, attr) and getattr(d, attr):
                setattr(survivor, attr, getattr(d, attr))
    if fb:
        survivor.product_feedback = fb[:50]
    await db.execute(sa.delete(SalesLead).where(SalesLead.id.in_(dup_ids)))
    await db.flush()
    return {"merged": len(dups), "moved_activities": moved_acts, "moved_wecom": moved_wecom}


async def import_from_excel(db: AsyncSession, *, content: bytes, source: str = "import") -> dict:
    """解析 .xlsx(首行表头:名称/电话/城市/行业/来源说明,列名容忍)→ 复用 import_leads。"""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "skipped": 0}
    header = [str(c or "").strip() for c in rows[0]]

    def _col(*names) -> int | None:
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None

    ci = {"name": _col("名称", "商家", "机构", "name"), "phone": _col("电话", "手机", "phone"),
          "city": _col("城市", "地区", "city"), "industry": _col("行业", "industry"),
          "note": _col("来源", "依据", "备注")}
    items: list[dict] = []
    for row in rows[1:]:
        def g(key):
            i = ci[key]
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else None
        name = g("name")
        if not name:
            continue
        items.append({"name": name, "phone": g("phone"), "region_name": g("city"),
                      "industry": g("industry"), "source_note": g("note")})
    if not items:
        return {"created": 0, "skipped": 0}
    return await import_leads(db, items=items, source=source)


# ── 座席权限 / 话术库 / 导出 ──────────────────────────────────────────────────

async def seat_scope_for(db: AsyncSession, admin_id: uuid.UUID) -> uuid.UUID | None:
    """若该 admin 是「座席」(在 seat_only_admin_ids 名单)→ 返回其 id(用于限定范围);否则 None(看全部)。"""
    cfg = await get_config(db)
    ids = {str(x) for x in (cfg.get("seat_only_admin_ids") or [])}
    return admin_id if str(admin_id) in ids else None


_SCRIPTS_KEY = "sales_scripts"


async def get_scripts(db: AsyncSession) -> list[dict]:
    """话术库/SOP:[{title, content, stage}]。存 system_configs.sales_scripts。"""
    row = (await db.execute(
        sa.select(SystemConfig).where(SystemConfig.key == _SCRIPTS_KEY))).scalar_one_or_none()
    if row is not None and isinstance(row.value, list):
        return row.value
    return []


async def set_scripts(db: AsyncSession, *, scripts: list[dict], updated_by: uuid.UUID) -> list[dict]:
    clean = [{"title": str(s.get("title", "")).strip(),
              "content": str(s.get("content", "")).strip(),
              "stage": s.get("stage") or None}
             for s in (scripts or []) if str(s.get("title", "")).strip()]
    row = (await db.execute(
        sa.select(SystemConfig).where(SystemConfig.key == _SCRIPTS_KEY))).scalar_one_or_none()
    if row is None:
        db.add(SystemConfig(id=uuid.uuid4(), key=_SCRIPTS_KEY, value=clean,
                            description="电销话术库 / 跟进 SOP", updated_by=updated_by))
    else:
        row.value = clean
        row.updated_by = updated_by
    await db.flush()
    return clean


_EXPORT_COLS = [
    ("name", "商家"), ("contact_name", "联系人"), ("phone", "电话"),
    ("region_name", "地区"), ("industry", "行业"), ("source", "来源"),
    ("status", "状态"), ("intent_grade", "意向"), ("intent_score", "意向分"),
    ("next_follow_at", "下次跟进"), ("source_note", "来源依据"),
]


async def export_leads_xlsx(db: AsyncSession, **filters) -> bytes:
    """按筛选导出线索为 .xlsx(最多 5000 条,超出请缩小筛选)。"""
    import io
    from openpyxl import Workbook
    filters.pop("skip", None)
    filters.pop("limit", None)
    rows, _total = await list_leads(db, skip=0, limit=5000, **filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "线索"
    ws.append([label for _k, label in _EXPORT_COLS])
    for r in rows:
        line = []
        for key, _label in _EXPORT_COLS:
            v = getattr(r, key, None)
            if key == "next_follow_at" and v is not None:
                v = v.strftime("%Y-%m-%d %H:%M")
            line.append("" if v is None else str(v))
        ws.append(line)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
